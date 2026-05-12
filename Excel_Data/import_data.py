"""
Import noodle delivery data from Excel into the NoodleDelivery app via its API.

Sources:
  - Copy of [Tháng 11-2025].xlsx  → November 2025 historical orders (ФОБО sheet)
  - Copy of 2026 [Huệ_Tina] Gửi tài xế hàng ngày.xlsx → current driver/restaurant roster

Pricing (per restaurant, per kg, Pho and Bun same price):
  - Thai chain restaurants  → 160 RUB/kg
  - All others              → 150 RUB/kg

Run with:
  python import_data.py
(The app must be running on http://localhost:5173 or adjust BASE_URL below)
"""

import sys, datetime, json
import urllib.request, urllib.error
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE_URL = "http://localhost:5035/api/v1"

EXCEL_DIR = "c:/Users/mattl/Downloads/Storm Forge Final Showcase 2026/Excel_Data"
FILE_2025 = f"{EXCEL_DIR}/Copy of [Tháng 11-2025].xlsx"
FILE_2026 = f"{EXCEL_DIR}/Copy of 2026 [Huệ_Tina] Gửi tài xế hàng ngày.xlsx"

# Thai chain restaurants get 160 RUB/kg; identified from 'Tính tiền Thái' sheet
THAI_KEYWORDS = ["kuzminki", "colombus", "pavele", "afimall", "salarit", "метропол", "thai"]

# Driver name normalization — map Excel names to canonical names
DRIVER_NAME_MAP = {
    "ali": "Ali",
    "али": "Ali",
    "yusup": "Yusup",
    "юсуф": "Yusup",
    "mukriddin": "Mukriddin",
    "мухриддин": "Mukriddin",
    "aleks/dima": "AlexDima",
    "alexdima": "AlexDima",
    "александр/дмитрий": "AlexDima",
    "владимир": "Vladimir",
    "vladimir": "Vladimir",
    "lan": "Lan",
    "ánh": "Anh",
    "anh": "Anh",
    "tuấn": "Tuan",
    "tuan": "Tuan",
    "tuấn tình": "Tuan Tinh",
    "ismail": "Ismail",
}


def normalize_driver(name: str) -> str:
    if not name:
        return name
    return DRIVER_NAME_MAP.get(name.strip().lower(), name.strip())


def is_thai(restaurant_name: str) -> bool:
    low = restaurant_name.lower()
    return any(kw in low for kw in THAI_KEYWORDS)


def api_get(path: str):
    url = f"{BASE_URL}{path}"
    with urllib.request.urlopen(url) as r:
        return json.loads(r.read())


def api_post(path: str, body: dict):
    url = f"{BASE_URL}{path}"
    data = json.dumps(body, ensure_ascii=False, default=str).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read())


def api_put(path: str, body: dict):
    url = f"{BASE_URL}{path}"
    data = json.dumps(body, ensure_ascii=False, default=str).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"}, method="PUT")
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read())


# ─────────────────────────────────────────────
# Step 1: Extract drivers from 2026 daily file
# ─────────────────────────────────────────────
def extract_drivers_2026():
    wb = openpyxl.load_workbook(FILE_2026, data_only=True)
    skip_sheets = {"tech", "Đếm Túi Ali", "Ghi Chú", "Kiểm lại", "Thông tin Xe giao hàng", "test"}
    drivers = []
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        name_cell = ws.cell(row=1, column=1).value
        phone_cell = ws.cell(row=1, column=2).value
        if not name_cell or not isinstance(name_cell, str):
            continue
        canonical = normalize_driver(name_cell)
        drivers.append({
            "raw": name_cell,
            "name": canonical,
            "phone": str(phone_cell).strip() if phone_cell else None,
        })
    return drivers


# ─────────────────────────────────────────────────────────────────────
# Step 2: Extract restaurants + orders from ФОБО sheet (Nov 2025)
# ─────────────────────────────────────────────────────────────────────
def extract_fbo_data():
    wb = openpyxl.load_workbook(FILE_2025, data_only=True)
    ws = wb["ФОБО (Phở Bò)"]

    # Row 1: headers — col 5 = ИТОГО, col 6..35 = days 1..30
    # Day columns: col 6 = day 1, col 7 = day 2, ... col 35 = day 30
    # (col 21 has a mid-month subtotal — skip it based on header value)
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Find which columns are actual day numbers (1-30)
    day_cols = {}  # day_number → column_index
    for c_idx, val in enumerate(header_row, start=1):
        if isinstance(val, (int, float)) and 1 <= val <= 31:
            day_cols[int(val)] = c_idx

    # Parse restaurant rows — pairs of (pho row, bun row)
    restaurants_data = []
    r = 3  # start after header rows
    while r <= ws.max_row:
        name = ws.cell(row=r, column=1).value
        product1 = ws.cell(row=r, column=2).value  # should be 'фо'
        driver = ws.cell(row=r, column=3).value

        if not name or name == "ИТОГО":
            r += 1
            continue

        # Pho quantities
        pho_by_day = {}
        for day, col in day_cols.items():
            v = ws.cell(row=r, column=col).value
            pho_by_day[day] = float(v) if isinstance(v, (int, float)) and v > 0 else 0.0

        # Next row is bun
        bun_by_day = {}
        if r + 1 <= ws.max_row:
            bun_product = ws.cell(row=r + 1, column=2).value
            if bun_product and "бун" in str(bun_product).lower():
                for day, col in day_cols.items():
                    v = ws.cell(row=r + 1, column=col).value
                    bun_by_day[day] = float(v) if isinstance(v, (int, float)) and v > 0 else 0.0
                r += 1  # consumed the bun row

        price = 160.0 if is_thai(name) else 150.0

        restaurants_data.append({
            "name": name.strip(),
            "driver": normalize_driver(driver) if driver else None,
            "price_per_kg": price,
            "pho_by_day": pho_by_day,
            "bun_by_day": bun_by_day,
        })
        r += 1

    return restaurants_data


# ────────────────────────────────────────────
# Step 3: Push drivers to API, return id map
# ────────────────────────────────────────────
def ensure_drivers(driver_list, restaurant_data):
    # Also collect driver names referenced in the ФОБО data
    fbo_drivers = set(r["driver"] for r in restaurant_data if r["driver"])
    all_drivers = {d["name"]: d.get("phone") for d in driver_list}
    for name in fbo_drivers:
        if name not in all_drivers:
            all_drivers[name] = None

    existing = {d["name"]: d["driver_id"] for d in api_get("/drivers?include_inactive=true")}
    driver_id_map = {}
    for name, phone in all_drivers.items():
        if name in existing:
            driver_id_map[name] = existing[name]
            print(f"  Driver exists: {name} (id={existing[name]})")
        else:
            status, resp = api_post("/drivers", {"name": name, "phone": phone})
            if status in (200, 201):
                driver_id_map[name] = resp["driver_id"]
                print(f"  Driver created: {name} (id={resp['driver_id']})")
            else:
                print(f"  ERROR creating driver {name}: {resp}")
    return driver_id_map


# ──────────────────────────────────────────────────
# Step 4: Push restaurants to API, return id map
# ──────────────────────────────────────────────────
def ensure_restaurants(restaurant_data):
    existing = {r["name"]: r for r in api_get("/restaurants?include_inactive=true")}
    restaurant_id_map = {}
    for r in restaurant_data:
        name = r["name"]
        if name in existing:
            rid = existing[name]["restaurant_id"]
            restaurant_id_map[name] = rid
            existing_price = existing[name].get("price_per_kg", 0)
            if abs(float(existing_price) - r["price_per_kg"]) > 0.01:
                api_put(f"/restaurants/{rid}", {
                    "name": name,
                    "address": existing[name].get("address"),
                    "phone": existing[name].get("phone"),
                    "is_active": True,
                    "price_per_kg": r["price_per_kg"],
                })
                print(f"  Restaurant updated price: {name} → {r['price_per_kg']} RUB/kg")
            else:
                print(f"  Restaurant exists: {name} (id={rid})")
        else:
            status, resp = api_post("/restaurants", {
                "name": name,
                "price_per_kg": r["price_per_kg"],
            })
            if status in (200, 201):
                restaurant_id_map[name] = resp["restaurant_id"]
                print(f"  Restaurant created: {name} (id={resp['restaurant_id']}, {r['price_per_kg']} RUB/kg)")
            else:
                print(f"  ERROR creating restaurant {name}: {resp}")
    return restaurant_id_map


# ────────────────────────────────────────────────────────────────
# Step 5: Submit orders day by day for November 2025
# ────────────────────────────────────────────────────────────────
def submit_orders(restaurant_data, driver_id_map, restaurant_id_map):
    # Get product IDs
    products = api_get("/products")
    pho_id = next((p["product_id"] for p in products if p["name"] == "Pho"), None)
    bun_id = next((p["product_id"] for p in products if p["name"] == "Bun"), None)
    if not pho_id or not bun_id:
        print("ERROR: Could not find Pho or Bun product IDs")
        return

    print(f"\n  Products: Pho id={pho_id}, Bun id={bun_id}")

    # Group by (date, driver) → list of (restaurant_id, pho_kg, bun_kg)
    # day_driver_orders[day][driver_name] = list of order entries
    day_driver_orders = {}

    for rest in restaurant_data:
        rest_name = rest["name"]
        driver_name = rest["driver"]
        rest_id = restaurant_id_map.get(rest_name)
        driver_id = driver_id_map.get(driver_name) if driver_name else None

        if not rest_id:
            print(f"  SKIP (no restaurant id): {rest_name}")
            continue
        if not driver_id:
            print(f"  SKIP (no driver): {rest_name} / driver={driver_name}")
            continue

        for day in range(1, 31):
            pho_kg = rest["pho_by_day"].get(day, 0.0)
            bun_kg = rest["bun_by_day"].get(day, 0.0)
            if pho_kg == 0 and bun_kg == 0:
                continue

            date_str = f"2025-11-{day:02d}"
            if date_str not in day_driver_orders:
                day_driver_orders[date_str] = {}
            if driver_id not in day_driver_orders[date_str]:
                day_driver_orders[date_str][driver_id] = []

            items = []
            if pho_kg > 0:
                items.append({"product_id": pho_id, "quantity_kg": pho_kg})
            if bun_kg > 0:
                items.append({"product_id": bun_id, "quantity_kg": bun_kg})

            day_driver_orders[date_str][driver_id].append({
                "driver_id": driver_id,
                "restaurant_id": rest_id,
                "items": items,
            })

    # Submit one batch per (date, driver)
    total_ok = 0
    total_skip = 0
    total_err = 0

    for date_str in sorted(day_driver_orders.keys()):
        for driver_id, order_entries in day_driver_orders[date_str].items():
            payload = {
                "date": date_str,
                "orders": order_entries,
            }
            status, resp = api_post("/orders/batch", payload)
            if status in (200, 201):
                total_ok += 1
            elif status == 409 and resp.get("code") == "DUPLICATE_ORDER":
                total_skip += 1  # already imported
            else:
                total_err += 1
                print(f"  ERROR {date_str} driver={driver_id}: {status} {resp}")

    print(f"\n  Orders submitted: {total_ok} ok, {total_skip} skipped (duplicates), {total_err} errors")


# ─────────
# Main
# ─────────
def main():
    print("=== NoodleDelivery Data Import ===\n")

    # Test connection
    try:
        api_get("/products")
    except Exception as e:
        print(f"Cannot reach API at {BASE_URL}\nMake sure the app is running.\nError: {e}")
        sys.exit(1)

    print("Step 1: Extracting drivers from 2026 file...")
    drivers = extract_drivers_2026()
    print(f"  Found {len(drivers)} driver sheets: {[d['name'] for d in drivers]}")

    print("\nStep 2: Extracting restaurant + order data from Nov 2025 ФОБО sheet...")
    restaurant_data = extract_fbo_data()
    print(f"  Found {len(restaurant_data)} restaurants")

    print("\nStep 3: Creating drivers...")
    driver_id_map = ensure_drivers(drivers, restaurant_data)

    print("\nStep 4: Creating restaurants...")
    restaurant_id_map = ensure_restaurants(restaurant_data)

    print("\nStep 5: Submitting November 2025 orders...")
    submit_orders(restaurant_data, driver_id_map, restaurant_id_map)

    print("\nDone!")


if __name__ == "__main__":
    main()
